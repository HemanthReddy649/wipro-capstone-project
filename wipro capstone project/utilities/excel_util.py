import openpyxl


def get_excel_data(file_path, sheet_name):

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    data = []

    for row in range(2, sheet.max_row + 1):

        row_data = {
            "departure": sheet.cell(row=row, column=1).value,
            "arrival": sheet.cell(row=row, column=2).value,
            "firstname": sheet.cell(row=row, column=3).value,
            "lastname": sheet.cell(row=row, column=4).value,
            "email": sheet.cell(row=row, column=5).value,
            "phone": sheet.cell(row=row, column=6).value
        }

        data.append(row_data)

    return data